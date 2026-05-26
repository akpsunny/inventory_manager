"""
core/database_manager.py
────────────────────────
SQLite-backed inventory store. Replaces the Master Inventory Excel file.

Why SQLite?
  • Bundled with Python (no extra dependency, no separate install)
  • Single-file database (.db) — easy to back up, email, version-control
  • ACID transactions — safe against partial-write corruption
  • Indexed lookups — faster than scanning Excel rows as the catalogue grows
  • Fully standalone — no server, no daemon, no separate process

The Current Stock REPORT is still generated as a styled .xlsx file
because that is a human deliverable, not the storage layer.
"""

from __future__ import annotations

import logging
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.fuzzy_matcher import FuzzyMatcher, THRESHOLD_MEDIUM

log = logging.getLogger(__name__)

SCHEMA_VERSION = 1

# ─── SQL Schema ───────────────────────────────────────────────────────────────
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS items (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    item_name     TEXT    NOT NULL,
    sku_code      TEXT    DEFAULT '',
    hsn_sac       TEXT    DEFAULT '',
    gst_rate      TEXT    DEFAULT '',
    rate          REAL    DEFAULT 0,
    measurement   TEXT    DEFAULT '',
    created_at    TEXT    DEFAULT CURRENT_TIMESTAMP,
    updated_at    TEXT    DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_items_sku  ON items(sku_code);
CREATE INDEX IF NOT EXISTS idx_items_name ON items(item_name);

CREATE TABLE IF NOT EXISTS invoice_receipts (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id       INTEGER NOT NULL,
    receipt_date  TEXT    NOT NULL,        -- DD/MM/YYYY
    quantity      REAL    NOT NULL DEFAULT 0,
    source_file   TEXT    DEFAULT '',
    created_at    TEXT    DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY(item_id) REFERENCES items(id) ON DELETE CASCADE,
    UNIQUE(item_id, receipt_date)
);
CREATE INDEX IF NOT EXISTS idx_receipts_item ON invoice_receipts(item_id);
CREATE INDEX IF NOT EXISTS idx_receipts_date ON invoice_receipts(receipt_date);

CREATE TABLE IF NOT EXISTS consumption (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id       INTEGER NOT NULL,
    quantity      REAL    NOT NULL DEFAULT 0,
    source_file   TEXT    DEFAULT '',
    recorded_at   TEXT    DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY(item_id) REFERENCES items(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_consumption_item ON consumption(item_id);

CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS processed_invoices (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_number  TEXT    DEFAULT '',     -- e.g. 'PF2511DL-0000876'; may be ''
    invoice_date    TEXT    DEFAULT '',     -- DD/MM/YYYY
    file_name       TEXT    DEFAULT '',     -- basename only, not full path
    items_count     INTEGER DEFAULT 0,
    processed_at    TEXT    DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_proc_inv_lookup
    ON processed_invoices(invoice_number, invoice_date);
CREATE INDEX IF NOT EXISTS idx_proc_inv_file
    ON processed_invoices(file_name, invoice_date);

CREATE TABLE IF NOT EXISTS manual_adjustments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id     INTEGER NOT NULL,
    delta_qty   REAL    NOT NULL,            -- positive = found stock; negative = damaged/missing
    note        TEXT    DEFAULT '',
    adjusted_at TEXT    DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY(item_id) REFERENCES items(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_adjustments_item ON manual_adjustments(item_id);
"""

# ─── Styling (for the Excel STOCK REPORT only) ────────────────────────────────
_GREEN_FILL = PatternFill("solid", fgColor="1A6B3C")
_DATE_FILL  = PatternFill("solid", fgColor="2E4057")
_CONS_FILL  = PatternFill("solid", fgColor="7B2D00")
_AVAIL_FILL = PatternFill("solid", fgColor="0B4F26")
_ALT_FILL   = PatternFill("solid", fgColor="F2F8F5")
_WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_BODY_FONT  = Font(name="Calibri", size=10)
_BOLD_FONT  = Font(name="Calibri", size=10, bold=True)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN_SIDE  = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE,  bottom=_THIN_SIDE)


def _safe_float(value) -> float:
    """Convert any value to float without raising. Handles ' 9 18 ', None, ''."""
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        for tok in str(value).split():
            try:
                return float(re.sub(r"[^\d.\-]", "", tok))
            except ValueError:
                continue
        return 0.0


# ─── DatabaseManager ──────────────────────────────────────────────────────────

class DatabaseManager:
    """
    SQLite-backed inventory store — drop-in replacement for ExcelManager.

    Public API (unchanged signatures from ExcelManager):
      create_empty_master(path)                       → str (final .db path)
      update_with_invoice(db_path, items, date_str)   → stats dict
      record_consumption(db_path, consumed_items)     → (matched, unmatched)
      generate_stock_report(db_path, out_path)        → stats dict

    Extra:
      migrate_from_excel(xlsx_path, db_path)          → stats dict
    """

    LOW_STOCK_THRESHOLD = 5

    def __init__(self):
        self._matcher = FuzzyMatcher(threshold=THRESHOLD_MEDIUM)

    # ── Connection ─────────────────────────────────────────────────────────

    def _connect(self, db_path: str) -> sqlite3.Connection:
        p = Path(db_path)
        # Catch a common mistake: pointing at a legacy Excel master directly.
        if p.suffix.lower() in {".xlsx", ".xls"}:
            raise ValueError(
                f"'{p.name}' is an Excel file, not a SQLite database. "
                "Use 'Migrate from Excel' in the Configuration section to "
                "convert it to a .db file, then select the new .db here."
            )
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("PRAGMA journal_mode = WAL")
            return conn
        except sqlite3.DatabaseError as e:
            raise ValueError(
                f"'{p.name}' is not a valid SQLite database ({e}). "
                "If this is a legacy Excel master, use 'Migrate from Excel' "
                "to convert it; otherwise create a fresh one via "
                "'Create New Master'."
            ) from e

    def _ensure_schema(self, conn: sqlite3.Connection):
        conn.executescript(SCHEMA_SQL)

    # ── Create new master ──────────────────────────────────────────────────

    def create_empty_master(self, path: str) -> str:
        """Initialise an empty SQLite DB with the schema. Returns final path."""
        p = Path(path)
        if p.suffix.lower() != ".db":
            p = p.with_suffix(".db")
        with self._connect(str(p)) as conn:
            self._ensure_schema(conn)
            conn.execute(
                "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                ("schema_version", str(SCHEMA_VERSION)))
            conn.execute(
                "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                ("created_at", datetime.now().isoformat()))
            conn.commit()
        log.info("Created new database: %s", p)
        return str(p)

    # ── Reset database ─────────────────────────────────────────────────────

    def reset_database(self, db_path: str) -> dict:
        """
        Delete all data from items, invoice_receipts, consumption,
        processed_invoices, and manual_adjustments. Keeps the database file,
        the schema, and the `meta` table intact, so the user's selected
        master path remains valid. Autoincrement counters are also reset so
        the next inserts start from id=1.

        Returns
        -------
        dict with the row counts that existed BEFORE the reset:
          {"items": int, "receipts": int, "consumption": int,
           "processed_invoices": int, "adjustments": int}
        """
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            counts = {
                "items":              conn.execute("SELECT COUNT(*) FROM items").fetchone()[0],
                "receipts":           conn.execute("SELECT COUNT(*) FROM invoice_receipts").fetchone()[0],
                "consumption":        conn.execute("SELECT COUNT(*) FROM consumption").fetchone()[0],
                "processed_invoices": conn.execute("SELECT COUNT(*) FROM processed_invoices").fetchone()[0],
                "adjustments":        conn.execute("SELECT COUNT(*) FROM manual_adjustments").fetchone()[0],
            }
            # Order matters even with FK cascade off — children first.
            conn.execute("DELETE FROM manual_adjustments")
            conn.execute("DELETE FROM consumption")
            conn.execute("DELETE FROM invoice_receipts")
            conn.execute("DELETE FROM items")
            # processed_invoices has no FK to items; clearing it lets the user
            # re-process the same invoice after a deliberate full reset.
            conn.execute("DELETE FROM processed_invoices")
            conn.execute(
                "DELETE FROM sqlite_sequence "
                "WHERE name IN ('items','invoice_receipts','consumption',"
                "'processed_invoices','manual_adjustments')")
            conn.execute(
                "INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                ("reset_at", datetime.now().isoformat()))
            conn.commit()
        log.info("Database reset: %s | cleared %s", db_path, counts)
        return counts

    # ── Inventory snapshot (for in-app data browser) ───────────────────────

    def get_inventory_snapshot(self, db_path: str) -> list[dict]:
        """
        Return every item with its computed totals, suitable for displaying in
        a read-only table in the UI. One dict per item with keys:
          id, item_name, sku_code, hsn_sac, gst_rate, rate, measurement,
          total_received, total_consumed, adjustments, available
        where available = total_received - total_consumed + adjustments.
        adjustments is the algebraic sum of manual stock corrections.
        """
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            rows = conn.execute("""
                SELECT
                  i.id, i.item_name, i.sku_code, i.hsn_sac, i.gst_rate,
                  i.rate, i.measurement,
                  COALESCE((SELECT SUM(quantity)  FROM invoice_receipts
                            WHERE item_id = i.id), 0) AS total_received,
                  COALESCE((SELECT SUM(quantity)  FROM consumption
                            WHERE item_id = i.id), 0) AS total_consumed,
                  COALESCE((SELECT SUM(delta_qty) FROM manual_adjustments
                            WHERE item_id = i.id), 0) AS total_adjustments
                FROM items i
                ORDER BY i.id
            """).fetchall()

        snapshot = []
        for r in rows:
            received    = float(r["total_received"]    or 0)
            consumed    = float(r["total_consumed"]    or 0)
            adjustments = float(r["total_adjustments"] or 0)
            snapshot.append({
                "id":             r["id"],
                "item_name":      r["item_name"] or "",
                "sku_code":       r["sku_code"]  or "",
                "hsn_sac":        r["hsn_sac"]   or "",
                "gst_rate":       r["gst_rate"]  or "",
                "rate":           float(r["rate"] or 0),
                "measurement":    r["measurement"] or "",
                "total_received": received,
                "total_consumed": consumed,
                "adjustments":    adjustments,
                "available":      max(0.0, received - consumed + adjustments),
            })
        return snapshot

    # ── Item CRUD (used by the data browser) ───────────────────────────────

    # Whitelist of columns the UI is allowed to update.  This guards against
    # accidentally writing to `id`, timestamps, or arbitrary names if the
    # caller passes a wider dict.
    _EDITABLE_ITEM_COLUMNS = {
        "item_name", "sku_code", "hsn_sac",
        "gst_rate", "rate", "measurement",
    }

    def update_item(self, db_path: str, item_id: int, **fields) -> None:
        """
        Update one item by id.  Only columns in _EDITABLE_ITEM_COLUMNS are
        respected; unknown keys are silently ignored.  Setting `item_name`
        to an empty value is rejected (every item must have a name).
        """
        clean = {k: v for k, v in fields.items()
                 if k in self._EDITABLE_ITEM_COLUMNS}
        if "item_name" in clean and not str(clean["item_name"]).strip():
            raise ValueError("Item name cannot be empty.")
        if "rate" in clean:
            clean["rate"] = max(0.0, float(clean["rate"] or 0))
        if not clean:
            return

        set_clause = ", ".join(f"{c} = ?" for c in clean)
        params     = list(clean.values()) + [
            datetime.now().isoformat(), int(item_id)]
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            cur = conn.execute(
                f"UPDATE items SET {set_clause}, updated_at = ? WHERE id = ?",
                params)
            if cur.rowcount == 0:
                raise ValueError(f"No item with id={item_id}")
            conn.commit()
        log.info("update_item id=%s fields=%s", item_id, list(clean))

    def delete_item(self, db_path: str, item_id: int) -> None:
        """
        Permanently delete an item and all its receipts, consumption rows,
        and manual adjustments (via FK cascade).  Irreversible — callers
        must confirm with the user first.
        """
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            cur = conn.execute("DELETE FROM items WHERE id = ?", (int(item_id),))
            if cur.rowcount == 0:
                raise ValueError(f"No item with id={item_id}")
            conn.commit()
        log.info("delete_item id=%s", item_id)

    def add_item(self, db_path: str, **fields) -> int:
        """
        Insert a new item.  Returns the new id.  item_name is required;
        everything else has a sensible default.
        """
        name = str(fields.get("item_name", "")).strip()
        if not name:
            raise ValueError("Item name is required.")
        sku  = str(fields.get("sku_code", "")).strip()
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            cur = conn.execute(
                "INSERT INTO items "
                "(item_name, sku_code, hsn_sac, gst_rate, rate, measurement) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (name, sku,
                 str(fields.get("hsn_sac", "")).strip(),
                 str(fields.get("gst_rate", "")).strip(),
                 max(0.0, float(fields.get("rate", 0) or 0)),
                 str(fields.get("measurement", "")).strip()))
            new_id = cur.lastrowid
            conn.commit()
        log.info("add_item id=%s name=%r sku=%r", new_id, name, sku)
        return new_id

    def add_manual_adjustment(
        self, db_path: str, item_id: int, delta_qty: float, note: str = ""
    ) -> None:
        """
        Append a signed quantity adjustment to manual_adjustments.
        Positive delta = stock found / correction up.
        Negative delta = damaged / lost / correction down.
        Each call appends a new row, so the full history is preserved and
        auditable.  Available qty is recomputed from the sum.
        """
        if abs(float(delta_qty)) < 1e-9:
            return   # zero-delta is a no-op, not an error
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            exists = conn.execute(
                "SELECT 1 FROM items WHERE id = ?", (int(item_id),)
            ).fetchone()
            if not exists:
                raise ValueError(f"No item with id={item_id}")
            conn.execute(
                "INSERT INTO manual_adjustments (item_id, delta_qty, note) "
                "VALUES (?, ?, ?)",
                (int(item_id), float(delta_qty), str(note or "").strip()))
            conn.commit()
        log.info("add_manual_adjustment id=%s delta=%+.2f note=%r",
                 item_id, delta_qty, note)

    # ── Duplicate-invoice detection ────────────────────────────────────────

    def find_processed_invoice(
        self,
        db_path:        str,
        invoice_number: str,
        invoice_date:   str,
        file_name:      str,
    ) -> Optional[dict]:
        """
        Check whether this invoice has already been logged. Match priority:
          1. If invoice_number is non-empty: match on (invoice_number, invoice_date).
          2. Otherwise fall back to (file_name, invoice_date).
        Returns the matching row as a dict (latest first by id), or None when
        no duplicate is found. Returns None unconditionally if invoice_date
        is empty — without a date we cannot dedupe reliably.
        """
        if not invoice_date:
            return None

        invoice_number = (invoice_number or "").strip()
        file_name      = (file_name or "").strip()

        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            if invoice_number:
                row = conn.execute(
                    "SELECT * FROM processed_invoices "
                    "WHERE invoice_number = ? AND invoice_date = ? "
                    "ORDER BY id DESC LIMIT 1",
                    (invoice_number, invoice_date)
                ).fetchone()
                if row:
                    return dict(row)
            if file_name:
                row = conn.execute(
                    "SELECT * FROM processed_invoices "
                    "WHERE file_name = ? AND invoice_date = ? "
                    "ORDER BY id DESC LIMIT 1",
                    (file_name, invoice_date)
                ).fetchone()
                if row:
                    return dict(row)
        return None

    def log_processed_invoice(
        self,
        db_path:        str,
        invoice_number: str,
        invoice_date:   str,
        file_name:      str,
        items_count:    int,
    ) -> None:
        """
        Record that an invoice was successfully processed.
        Always appends a new row — re-processed invoices show up multiple
        times with distinct processed_at timestamps, which is useful when
        auditing how a quantity total accumulated.
        """
        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            conn.execute(
                "INSERT INTO processed_invoices "
                "(invoice_number, invoice_date, file_name, items_count) "
                "VALUES (?, ?, ?, ?)",
                ((invoice_number or "").strip(),
                 (invoice_date   or "").strip(),
                 (file_name      or "").strip(),
                 int(items_count or 0))
            )
            conn.commit()
        log.info("Logged processed invoice: number=%r date=%r file=%r items=%d",
                 invoice_number, invoice_date, file_name, items_count)

    # ── Update with invoice ────────────────────────────────────────────────

    def update_with_invoice(
        self, db_path: str, items: list[dict], date_str: str,
    ) -> dict:
        """
        Merge invoice line items into the database.
          • Existing item (by SKU/name, exact or fuzzy) → add qty to that date.
          • New item → INSERT into items, then INSERT into invoice_receipts.
          • Same item + same date called twice → quantities accumulate via UPSERT.
        """
        stats = {"new_rows": 0, "updated_rows": 0, "skipped": 0}

        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            sku_index, name_index = self._build_indexes(conn)

            for item in items:
                sku       = str(item.get("sku_code", "")).strip()
                item_name = str(item.get("item_name", "")).strip()
                qty       = _safe_float(item.get("quantity", 0))

                if not item_name:
                    stats["skipped"] += 1
                    continue

                item_id = self._find_item_id(sku, item_name,
                                             sku_index, name_index,
                                             strict_sku=True)

                if item_id is None:
                    item_id = self._insert_item(conn, item)
                    if sku:
                        sku_index[sku.lower()] = item_id
                    name_index[item_name.lower()] = item_id
                    stats["new_rows"] += 1
                else:
                    stats["updated_rows"] += 1

                # UPSERT into invoice_receipts (UNIQUE on item_id+date)
                conn.execute("""
                    INSERT INTO invoice_receipts (item_id, receipt_date, quantity)
                    VALUES (?, ?, ?)
                    ON CONFLICT(item_id, receipt_date)
                    DO UPDATE SET quantity = quantity + excluded.quantity
                """, (item_id, date_str, qty))

            conn.commit()

        log.info("DB updated: %s", stats)
        return stats

    # ── Record consumption ─────────────────────────────────────────────────

    def record_consumption(
        self, db_path: str, consumed_items: list[dict],
    ) -> tuple[int, int]:
        """
        Append rows to the consumption table. Items that cannot be matched
        to the master are reported as unmatched (not auto-created).
        """
        matched = unmatched = 0

        with self._connect(db_path) as conn:
            self._ensure_schema(conn)
            sku_index, name_index = self._build_indexes(conn)

            for item in consumed_items:
                sku       = str(item.get("sku_code", "")).strip()
                item_name = str(item.get("item_name", "")).strip()
                qty       = _safe_float(item.get("quantity", 0))

                item_id = self._find_item_id(sku, item_name,
                                             sku_index, name_index)
                if item_id is None:
                    log.warning("Unmatched consumption: '%s' (SKU '%s')",
                                item_name, sku)
                    unmatched += 1
                    continue

                conn.execute(
                    "INSERT INTO consumption (item_id, quantity) VALUES (?, ?)",
                    (item_id, qty))
                matched += 1

            conn.commit()

        return matched, unmatched

    # ── Generate stock report (Excel output) ───────────────────────────────

    def generate_stock_report(self, db_path: str, out_path: str) -> dict:
        with self._connect(db_path) as conn:
            items_rows = conn.execute(
                "SELECT * FROM items ORDER BY id").fetchall()

            received = {r["item_id"]: r["total"] for r in conn.execute(
                "SELECT item_id, SUM(quantity) AS total "
                "FROM invoice_receipts GROUP BY item_id").fetchall()}

            consumed = {r["item_id"]: r["total"] for r in conn.execute(
                "SELECT item_id, SUM(quantity) AS total "
                "FROM consumption GROUP BY item_id").fetchall()}

            distinct_dates = conn.execute(
                "SELECT COUNT(DISTINCT receipt_date) AS n "
                "FROM invoice_receipts").fetchone()["n"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Current Stock"

        headers = ["S.No.", "Item Name", "SKU Code", "HSN/SAC",
                   "GST Rate", "Rate", "Measurement",
                   "Total Received", "Consumed Qty", "Available Qty", "Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            if h == "Available Qty":
                self._style_header(cell, _AVAIL_FILL)
            elif h == "Consumed Qty":
                self._style_header(cell, _CONS_FILL)
            elif h == "Status":
                self._style_header(cell, _DATE_FILL)
            else:
                self._style_header(cell)

        stats = {"total_skus": 0, "date_columns": distinct_dates,
                 "low_stock": 0, "zero_stock": 0}

        out_row = 2
        for i, item in enumerate(items_rows, 1):
            total_received = float(received.get(item["id"], 0) or 0)
            total_consumed = float(consumed.get(item["id"], 0) or 0)
            available = max(0.0, total_received - total_consumed)

            values = [i, item["item_name"], item["sku_code"], item["hsn_sac"],
                      item["gst_rate"], item["rate"], item["measurement"],
                      round(total_received, 2), round(total_consumed, 2),
                      round(available, 2)]
            for col_idx, v in enumerate(values, 1):
                cell = ws.cell(row=out_row, column=col_idx, value=v)
                self._style_body(cell, alt=(out_row % 2 == 0),
                                 bold=(col_idx == 2))

            if available <= 0:
                status, color = "❌ Out of Stock", "C0392B"
                stats["zero_stock"] += 1
            elif available <= self.LOW_STOCK_THRESHOLD:
                status, color = "⚠ Low Stock", "E67E22"
                stats["low_stock"] += 1
            else:
                status, color = "✔ In Stock", "1A6B3C"

            scell = ws.cell(row=out_row, column=len(headers), value=status)
            scell.font = Font(color=color, bold=True, name="Calibri", size=10)
            scell.alignment = _CENTER
            scell.border = _THIN_BORDER
            if out_row % 2 == 0:
                scell.fill = _ALT_FILL

            stats["total_skus"] += 1
            out_row += 1

        widths = {1: 7, 2: 35, 3: 16, 4: 13, 5: 10, 6: 11, 7: 14,
                  8: 15, 9: 14, 10: 14, 11: 16}
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        self._add_summary_sheet(wb, stats, db_path, distinct_dates)
        wb.save(out_path)
        log.info("Stock report saved: %s | %s", out_path, stats)
        return stats

    # ── One-time migration from existing Excel master ──────────────────────

    def migrate_from_excel(self, xlsx_path: str, db_path: str) -> dict:
        """Import an existing Master_Inventory.xlsx into a fresh SQLite DB."""
        import openpyxl

        self.create_empty_master(db_path)

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"items": 0, "receipts": 0, "consumption": 0}

        headers = [str(c).strip() if c is not None else "" for c in rows[0]]

        def idx_of(h: str) -> Optional[int]:
            return headers.index(h) if h in headers else None

        idx = {
            "item_name":   idx_of("Item Name"),
            "sku_code":    idx_of("SKU Code"),
            "hsn_sac":     idx_of("HSN/SAC"),
            "gst_rate":    idx_of("GST Rate"),
            "rate":        idx_of("Rate"),
            "measurement": idx_of("Measurement"),
            "consumed":    idx_of("Consumed Qty"),
        }
        date_cols = [(i, h) for i, h in enumerate(headers)
                     if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", h or "")]

        stats = {"items": 0, "receipts": 0, "consumption": 0}

        def _cell(row, key):
            i = idx[key]
            return row[i] if i is not None and i < len(row) else None

        with self._connect(db_path) as conn:
            for row in rows[1:]:
                name = str(_cell(row, "item_name") or "").strip()
                if not name:
                    continue

                cur = conn.execute(
                    "INSERT INTO items "
                    "(item_name, sku_code, hsn_sac, gst_rate, rate, measurement) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (name,
                     str(_cell(row, "sku_code")    or "").strip(),
                     str(_cell(row, "hsn_sac")     or "").strip(),
                     str(_cell(row, "gst_rate")    or "").strip(),
                     _safe_float(_cell(row, "rate")),
                     str(_cell(row, "measurement") or "").strip()))
                item_id = cur.lastrowid
                stats["items"] += 1

                for col_idx, date_str in date_cols:
                    qty = _safe_float(row[col_idx] if col_idx < len(row) else 0)
                    if qty > 0:
                        conn.execute(
                            "INSERT INTO invoice_receipts "
                            "(item_id, receipt_date, quantity) VALUES (?, ?, ?)",
                            (item_id, date_str, qty))
                        stats["receipts"] += 1

                cqty = _safe_float(_cell(row, "consumed"))
                if cqty > 0:
                    conn.execute(
                        "INSERT INTO consumption "
                        "(item_id, quantity, source_file) VALUES (?, ?, ?)",
                        (item_id, cqty, f"migrated from {Path(xlsx_path).name}"))
                    stats["consumption"] += 1

            conn.commit()

        log.info("Migrated %s → %s | %s", xlsx_path, db_path, stats)
        return stats

    # ── Item lookup / insert helpers ───────────────────────────────────────

    def _build_indexes(self, conn) -> tuple[dict, dict]:
        sku_idx, name_idx = {}, {}
        for row in conn.execute("SELECT id, sku_code, item_name FROM items"):
            if row["sku_code"]:
                sku_idx[row["sku_code"].strip().lower()] = row["id"]
            if row["item_name"]:
                name_idx[row["item_name"].strip().lower()] = row["id"]
        return sku_idx, name_idx

    def _find_item_id(self, sku: str, item_name: str,
                      sku_index: dict, name_index: dict,
                      *, strict_sku: bool = False) -> Optional[int]:
        """
        Look up an existing item by SKU then by name.

        strict_sku=False (default, used by consumption matching):
          Tries exact SKU, fuzzy SKU (≥90), exact name, fuzzy name (≥78).
          Fuzzy fallback is needed when the user's consumption report has
          slightly different wording than the invoice.

        strict_sku=True (used by invoice ingestion):
          Only exact matches count. SKU path: exact SKU; if SKU given but no
          match → return None (the caller inserts a new item). Name path:
          exact name; if name given but no exact match → return None. No
          fuzzy step on either axis.

          Rationale: supplier-provided SKUs are authoritative identifiers.
          Fuzzy-merging SKUHDR-0001402 (Grey Ceramics) into SKUHDR-0001403
          (White Ceramics) collapses distinct products. The same applies to
          names — 'AC-143 Ceramic Pot' fuzzy-matches 'AB-184 Ceramic Pot' at
          ~89% but they are completely different products. When a row has no
          SKU (e.g. because the source PDF column was empty and we cleared a
          bogus value upstream), the right behaviour is to insert it as a
          new item, not absorb it into an unrelated lookalike.
        """
        sku_norm  = sku.strip().lower()
        name_norm = item_name.strip().lower()

        if sku_norm:
            if sku_norm in sku_index:
                return sku_index[sku_norm]
            if strict_sku:
                return None
            if sku_index:
                best, _ = self._matcher.best_match(
                    sku_norm, list(sku_index.keys()), threshold=90)
                if best:
                    return sku_index[best]

        if name_norm:
            if name_norm in name_index:
                return name_index[name_norm]
            if strict_sku:
                return None
            if name_index:
                best, score = self._matcher.best_match(
                    name_norm, list(name_index.keys()),
                    threshold=THRESHOLD_MEDIUM)
                if best:
                    log.debug("Fuzzy match '%s' → '%s' (%d)",
                              item_name, best, score)
                    return name_index[best]
        return None

    def _insert_item(self, conn, item: dict) -> int:
        cur = conn.execute(
            "INSERT INTO items "
            "(item_name, sku_code, hsn_sac, gst_rate, rate, measurement) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (str(item.get("item_name", "")).strip(),
             str(item.get("sku_code", "")).strip(),
             str(item.get("hsn_sac", "")).strip(),
             str(item.get("gst_rate", "")).strip(),
             _safe_float(item.get("rate", 0)),
             str(item.get("measurement", "")).strip()))
        return cur.lastrowid

    # ── Excel report styling helpers ───────────────────────────────────────

    @staticmethod
    def _style_header(cell, fill=None):
        cell.font, cell.alignment = _WHITE_FONT, _CENTER
        cell.border, cell.fill = _THIN_BORDER, fill or _GREEN_FILL

    @staticmethod
    def _style_body(cell, alt=False, bold=False):
        cell.font = _BOLD_FONT if bold else _BODY_FONT
        cell.alignment, cell.border = _CENTER, _THIN_BORDER
        if alt:
            cell.fill = _ALT_FILL

    def _add_summary_sheet(self, wb, stats, db_path, distinct_dates):
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False

        rows = [
            ("Generated On",     datetime.now().strftime("%d/%m/%Y  %H:%M:%S")),
            ("Database File",    str(db_path)),
            ("Total SKUs",       stats["total_skus"]),
            ("Distinct Dates",   distinct_dates),
            ("In Stock Items",   stats["total_skus"] - stats["low_stock"] - stats["zero_stock"]),
            ("Low Stock Items",  stats["low_stock"]),
            ("Out of Stock",     stats["zero_stock"]),
        ]
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 55

        for r_idx, (label, value) in enumerate(rows, 2):
            lc = ws.cell(row=r_idx, column=1, value=label)
            lc.font = Font(name="Calibri", size=11, bold=True, color="1A6B3C")
            lc.alignment = _LEFT
            vc = ws.cell(row=r_idx, column=2, value=value)
            vc.font = Font(name="Calibri", size=11)
            vc.alignment = _LEFT

        title = ws.cell(row=1, column=1,
                        value="📦  Current Stock Report — Summary")
        ws.merge_cells("A1:B1")
        title.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title.fill, title.alignment = _GREEN_FILL, _CENTER
        ws.row_dimensions[1].height = 36
