"""
core/excel_manager.py
─────────────────────
Manages all interactions with the Master Inventory Excel file and generates
the final Current Stock report.

Master file schema
──────────────────
  S.No. | Item Name | SKU Code | HSN/SAC | GST Rate | Rate | Measurement
  | [DD/MM/YYYY] | [DD/MM/YYYY] | … | Consumed Qty | Available Qty

  • Columns 1–7  : static item details (never change once set)
  • Columns 8+   : one column per invoice date  (quantities received)
  • Last-2       : Consumed Qty  (updated by consumption processor)
  • Last-1       : Available Qty (computed on report generation)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd

from core.fuzzy_matcher import FuzzyMatcher, THRESHOLD_MEDIUM, THRESHOLD_LOW

log = logging.getLogger(__name__)

# ─── Schema constants ─────────────────────────────────────────────────────────
STATIC_COLS = ["S.No.", "Item Name", "SKU Code", "HSN/SAC",
               "GST Rate", "Rate", "Measurement"]
CONSUMED_COL    = "Consumed Qty"
AVAILABLE_COL   = "Available Qty"
STATIC_COL_COUNT = len(STATIC_COLS)   # 7

# ─── Styling helpers ──────────────────────────────────────────────────────────
_GREEN_FILL   = PatternFill("solid", fgColor="1A6B3C")
_DATE_FILL    = PatternFill("solid", fgColor="2E4057")
_CONS_FILL    = PatternFill("solid", fgColor="7B2D00")
_AVAIL_FILL   = PatternFill("solid", fgColor="0B4F26")
_ALT_FILL     = PatternFill("solid", fgColor="F2F8F5")
_WHITE_FONT   = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_BODY_FONT    = Font(name="Calibri", size=10)
_BOLD_FONT    = Font(name="Calibri", size=10, bold=True)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN_SIDE    = Side(style="thin", color="CCCCCC")
_THIN_BORDER  = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                        top=_THIN_SIDE,  bottom=_THIN_SIDE)


def _style_header_cell(cell, fill=None):
    cell.font      = _WHITE_FONT
    cell.alignment = _CENTER
    cell.border    = _THIN_BORDER
    cell.fill      = fill or _GREEN_FILL


def _style_body_cell(cell, alt: bool = False, bold: bool = False):
    cell.font      = _BOLD_FONT if bold else _BODY_FONT
    cell.alignment = _CENTER
    cell.border    = _THIN_BORDER
    if alt:
        cell.fill  = _ALT_FILL


# ─── Utility ──────────────────────────────────────────────────────────────────

def _safe_float(value) -> float:
    """
    Convert any value to float without raising.
    Handles strings like '9 18' (takes first numeric token), None, '', etc.
    """
    if value is None or value == "":
        return 0.0
    text = str(value).strip()
    # Direct conversion
    try:
        return float(text)
    except ValueError:
        pass
    # Space-separated tokens → take first valid number
    import re as _re
    for token in text.split():
        try:
            return float(_re.sub(r"[^\d.\-]", "", token))
        except ValueError:
            continue
    return 0.0


def _is_date_col(header: str) -> bool:
    """Return True if the header looks like a DD/MM/YYYY date string."""
    return bool(re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", str(header).strip()))


def _col_letter(ws, header: str) -> Optional[str]:
    """Return the Excel column letter for a given header value, or None."""
    for cell in ws[1]:
        if str(cell.value).strip() == header:
            return get_column_letter(cell.column)
    return None


def _col_index(ws, header: str) -> Optional[int]:
    """Return the 1-based column index for a given header value, or None."""
    for cell in ws[1]:
        if str(cell.value or "").strip() == header:
            return cell.column
    return None


# ─── ExcelManager ─────────────────────────────────────────────────────────────

class ExcelManager:
    """
    All operations on the Master Inventory Excel file.

    Public methods
    ──────────────
    create_empty_master(path)
    update_with_invoice(master_path, items, date_str) → stats dict
    record_consumption(master_path, consumed_items)   → (matched, unmatched)
    generate_stock_report(master_path, out_path)      → stats dict
    """

    def __init__(self):
        self._matcher = FuzzyMatcher(threshold=THRESHOLD_MEDIUM)

    # ── Create new master ──────────────────────────────────────────────────

    def create_empty_master(self, path: str):
        """Create a blank Master Inventory Excel file with the correct schema."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Inventory"

        all_headers = STATIC_COLS + [CONSUMED_COL, AVAILABLE_COL]
        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if header == CONSUMED_COL:
                _style_header_cell(cell, _CONS_FILL)
            elif header == AVAILABLE_COL:
                _style_header_cell(cell, _AVAIL_FILL)
            else:
                _style_header_cell(cell)

        self._set_col_widths(ws, all_headers)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        wb.save(path)
        log.info("Created new master: %s", path)

    # ── Update with invoice ────────────────────────────────────────────────

    def update_with_invoice(
        self,
        master_path: str,
        items: list[dict],
        date_str: str,
    ) -> dict:
        """
        Merge invoice line items into the master file.

        For each item:
          • If SKU or name already exists → update quantity in date column.
          • If new                        → add a new row.

        Returns a stats dict: {new_rows, updated_rows, skipped}.
        """
        wb = openpyxl.load_workbook(master_path)
        ws = wb.active

        # Ensure the date column exists (inserts before Consumed/Available)
        date_col_idx = self._ensure_date_column(ws, date_str)

        # Build lookup index: {normalised_key: row_number}
        sku_index  = self._build_sku_index(ws)
        name_index = self._build_name_index(ws)

        stats = {"new_rows": 0, "updated_rows": 0, "skipped": 0}

        for item in items:
            sku       = str(item.get("sku_code", "")).strip()
            item_name = str(item.get("item_name", "")).strip()
            qty       = _safe_float(item.get("quantity", 0))

            if not item_name:
                stats["skipped"] += 1
                continue

            existing_row = self._find_row(sku, item_name, sku_index, name_index)

            if existing_row:
                # Update existing row
                current = ws.cell(row=existing_row, column=date_col_idx).value
                current_qty = _safe_float(current)
                ws.cell(row=existing_row, column=date_col_idx).value = current_qty + qty
                _style_body_cell(ws.cell(row=existing_row, column=date_col_idx),
                                 alt=(existing_row % 2 == 0))
                stats["updated_rows"] += 1
                log.debug("Updated row %d: %s += %s", existing_row, item_name, qty)
            else:
                # Insert new row (before trailing consumed/available cols)
                new_row = ws.max_row + 1
                sno = new_row - 1   # row 1 = header, so row 2 = S.No. 1
                self._write_new_row(ws, new_row, sno, item, date_col_idx, qty)
                # Refresh indices
                sku_index[sku.lower()]        = new_row
                name_index[item_name.lower()] = new_row
                stats["new_rows"] += 1
                log.debug("New row %d: %s", new_row, item_name)

        self._apply_body_styles(ws)
        wb.save(master_path)
        log.info("Master updated: %s", stats)
        return stats

    # ── Record consumption ─────────────────────────────────────────────────

    def record_consumption(
        self,
        master_path: str,
        consumed_items: list[dict],
    ) -> tuple[int, int]:
        """
        Add consumed quantities from a consumption report into the
        'Consumed Qty' column of the master file.

        Returns (matched_count, unmatched_count).
        """
        wb = openpyxl.load_workbook(master_path)
        ws = wb.active

        consumed_col_idx = _col_index(ws, CONSUMED_COL)
        if consumed_col_idx is None:
            # Insert it if missing
            consumed_col_idx = self._ensure_trailing_cols(ws)

        sku_index  = self._build_sku_index(ws)
        name_index = self._build_name_index(ws)

        matched   = 0
        unmatched = 0

        for item in consumed_items:
            sku       = str(item.get("sku_code", "")).strip()
            item_name = str(item.get("item_name", "")).strip()
            qty       = _safe_float(item.get("quantity", 0))

            row = self._find_row(sku, item_name, sku_index, name_index)

            if row:
                current = ws.cell(row=row, column=consumed_col_idx).value
                ws.cell(row=row, column=consumed_col_idx).value = _safe_float(current) + qty
                matched += 1
            else:
                log.warning("Unmatched consumption item: '%s' (SKU: '%s')",
                            item_name, sku)
                unmatched += 1

        self._apply_body_styles(ws)
        wb.save(master_path)
        return matched, unmatched

    # ── Generate stock report ──────────────────────────────────────────────

    def generate_stock_report(self, master_path: str, out_path: str) -> dict:
        """
        Read the master file, compute available stock for every item,
        and write a formatted Current Stock report Excel file.

        Available Qty = Σ(all date columns) − Consumed Qty
        """
        wb_master = openpyxl.load_workbook(master_path)
        ws_master = wb_master.active

        # Identify column indices
        headers = [ws_master.cell(1, c).value
                   for c in range(1, ws_master.max_column + 1)]
        date_col_indices = [
            i + 1 for i, h in enumerate(headers) if _is_date_col(str(h or ""))
        ]
        consumed_col_idx = next(
            (i + 1 for i, h in enumerate(headers) if str(h or "") == CONSUMED_COL),
            None)
        available_col_idx = next(
            (i + 1 for i, h in enumerate(headers) if str(h or "") == AVAILABLE_COL),
            None)

        LOW_STOCK_THRESHOLD = 5

        # Build output workbook
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Current Stock"

        # Write output headers
        out_headers = STATIC_COLS + ["Total Received", CONSUMED_COL,
                                     AVAILABLE_COL, "Status"]
        for col_idx, h in enumerate(out_headers, 1):
            cell = ws_out.cell(row=1, column=col_idx, value=h)
            if h == AVAILABLE_COL:
                _style_header_cell(cell, _AVAIL_FILL)
            elif h == CONSUMED_COL:
                _style_header_cell(cell, _CONS_FILL)
            elif h == "Status":
                _style_header_cell(cell, _DATE_FILL)
            else:
                _style_header_cell(cell)

        stats = {
            "total_skus": 0,
            "date_columns": len(date_col_indices),
            "low_stock": 0,
            "zero_stock": 0,
        }

        out_row = 2
        for row in range(2, ws_master.max_row + 1):
            item_name = ws_master.cell(row, 2).value
            if not item_name:
                continue

            # Sum all date columns
            total_received = sum(
                float(ws_master.cell(row, c).value or 0)
                for c in date_col_indices)

            consumed = _safe_float(
                ws_master.cell(row, consumed_col_idx).value
                if consumed_col_idx else 0)

            available = max(0.0, total_received - consumed)

            # Static columns
            for out_c, master_c in enumerate(range(1, STATIC_COL_COUNT + 1), 1):
                v = ws_master.cell(row, master_c).value
                cell = ws_out.cell(row=out_row, column=out_c, value=v)
                _style_body_cell(cell, alt=(out_row % 2 == 0),
                                  bold=(out_c == 2))

            # Computed columns
            col_offset = STATIC_COL_COUNT + 1
            for val in [total_received, consumed, available]:
                cell = ws_out.cell(row=out_row, column=col_offset, value=round(val, 2))
                _style_body_cell(cell, alt=(out_row % 2 == 0))
                col_offset += 1

            # Status
            if available <= 0:
                status = "❌ Out of Stock"
                status_font = Font(color="C0392B", bold=True, name="Calibri", size=10)
                stats["zero_stock"] += 1
            elif available <= LOW_STOCK_THRESHOLD:
                status = "⚠ Low Stock"
                status_font = Font(color="E67E22", bold=True, name="Calibri", size=10)
                stats["low_stock"] += 1
            else:
                status = "✔ In Stock"
                status_font = Font(color="1A6B3C", bold=True, name="Calibri", size=10)

            status_cell = ws_out.cell(row=out_row, column=col_offset, value=status)
            status_cell.font      = status_font
            status_cell.alignment = _CENTER
            status_cell.border    = _THIN_BORDER
            if out_row % 2 == 0:
                status_cell.fill = _ALT_FILL

            stats["total_skus"] += 1
            out_row += 1

        # Column widths & freeze
        self._set_col_widths(ws_out, out_headers)
        ws_out.row_dimensions[1].height = 32
        ws_out.freeze_panes = "A2"

        # Add a summary sheet
        self._add_summary_sheet(wb_out, stats,
                                master_path, len(date_col_indices))

        wb_out.save(out_path)
        log.info("Stock report saved: %s | Stats: %s", out_path, stats)
        return stats

    # ── Internal helpers ───────────────────────────────────────────────────

    def _ensure_date_column(self, ws, date_str: str) -> int:
        """
        Make sure a column with header=date_str exists.
        It is inserted just before the 'Consumed Qty' column.
        Returns the column index (1-based).
        """
        # Check if it already exists
        existing = _col_index(ws, date_str)
        if existing:
            return existing

        # Find insertion point: before Consumed Qty (or Available Qty)
        insert_before = _col_index(ws, CONSUMED_COL)
        if insert_before is None:
            insert_before = _col_index(ws, AVAILABLE_COL)
        if insert_before is None:
            insert_before = ws.max_column + 1

        ws.insert_cols(insert_before)
        new_cell = ws.cell(row=1, column=insert_before, value=date_str)
        _style_header_cell(new_cell, _DATE_FILL)
        ws.column_dimensions[get_column_letter(insert_before)].width = 14
        return insert_before

    def _ensure_trailing_cols(self, ws) -> int:
        """Append Consumed Qty and Available Qty columns if absent."""
        max_col = ws.max_column
        for label, fill in [(CONSUMED_COL, _CONS_FILL), (AVAILABLE_COL, _AVAIL_FILL)]:
            if _col_index(ws, label) is None:
                max_col += 1
                cell = ws.cell(row=1, column=max_col, value=label)
                _style_header_cell(cell, fill)
        return _col_index(ws, CONSUMED_COL)

    def _build_sku_index(self, ws) -> dict[str, int]:
        """Return {normalised_sku: row_number} for all data rows."""
        sku_col = _col_index(ws, "SKU Code")
        if not sku_col:
            return {}
        idx: dict[str, int] = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, sku_col).value
            if val:
                idx[str(val).strip().lower()] = row
        return idx

    def _build_name_index(self, ws) -> dict[str, int]:
        """Return {normalised_name: row_number} for all data rows."""
        name_col = _col_index(ws, "Item Name")
        if not name_col:
            return {}
        idx: dict[str, int] = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, name_col).value
            if val:
                idx[str(val).strip().lower()] = row
        return idx

    def _find_row(
        self,
        sku: str,
        item_name: str,
        sku_index: dict,
        name_index: dict,
    ) -> Optional[int]:
        """
        Locate the master-file row for a given item.
        Priority:
          1. Exact SKU match
          2. Fuzzy SKU match (≥90)
          3. Exact name match
          4. Fuzzy name match (≥78)
        """
        sku_norm  = sku.strip().lower()
        name_norm = item_name.strip().lower()

        # 1. Exact SKU
        if sku_norm and sku_norm in sku_index:
            return sku_index[sku_norm]

        # 2. Fuzzy SKU
        if sku_norm and sku_index:
            best_sku, score = self._matcher.best_match(sku_norm, list(sku_index.keys()),
                                                        threshold=90)
            if best_sku:
                return sku_index[best_sku]

        # 3. Exact name
        if name_norm and name_norm in name_index:
            return name_index[name_norm]

        # 4. Fuzzy name
        if name_norm and name_index:
            best_name, score = self._matcher.best_match(
                name_norm, list(name_index.keys()), threshold=THRESHOLD_MEDIUM)
            if best_name:
                log.debug("Fuzzy name match: '%s' → '%s' (score %d)",
                          item_name, best_name, score)
                return name_index[best_name]

        return None

    def _write_new_row(self, ws, row: int, sno: int, item: dict,
                       date_col_idx: int, qty: float):
        """Write all columns for a brand-new item row."""
        alt = row % 2 == 0

        values = {
            1: sno,
            2: item.get("item_name", ""),
            3: item.get("sku_code", ""),
            4: item.get("hsn_sac", ""),
            5: item.get("gst_rate", ""),
            6: _safe_float(item.get("rate", 0)),
            7: item.get("measurement", ""),
        }
        for col, val in values.items():
            cell = ws.cell(row=row, column=col, value=val)
            _style_body_cell(cell, alt=alt, bold=(col == 2))

        # Quantity in date column
        qty_cell = ws.cell(row=row, column=date_col_idx, value=qty)
        _style_body_cell(qty_cell, alt=alt)

    def _apply_body_styles(self, ws):
        """Ensure all data rows have consistent styling."""
        for row in range(2, ws.max_row + 1):
            alt = row % 2 == 0
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and not cell.has_style:
                    _style_body_cell(cell, alt=alt)
            ws.row_dimensions[row].height = 18

    def _set_col_widths(self, ws, headers: list[str]):
        """Set sensible column widths based on header names."""
        widths = {
            "S.No.":        7,
            "Item Name":    35,
            "SKU Code":     16,
            "HSN/SAC":      13,
            "GST Rate":     10,
            "Rate":         11,
            "Measurement":  14,
            CONSUMED_COL:   14,
            AVAILABLE_COL:  14,
            "Total Received": 15,
            "Status":       16,
        }
        for col_idx, h in enumerate(headers, 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = widths.get(str(h), 14)

    def _add_summary_sheet(self, wb: Workbook, stats: dict,
                            master_path: str, date_count: int):
        """Add a one-page summary sheet to the output workbook."""
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False

        rows = [
            ("Generated On",    datetime.now().strftime("%d/%m/%Y  %H:%M:%S")),
            ("Master File",     str(master_path)),
            ("Total SKUs",      stats["total_skus"]),
            ("Invoice Dates",   date_count),
            ("In Stock Items",  stats["total_skus"] - stats["low_stock"] - stats["zero_stock"]),
            ("Low Stock Items", stats["low_stock"]),
            ("Out of Stock",    stats["zero_stock"]),
        ]

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 55

        for r_idx, (label, value) in enumerate(rows, 2):
            label_cell = ws.cell(row=r_idx, column=1, value=label)
            label_cell.font      = Font(name="Calibri", size=11, bold=True, color="1A6B3C")
            label_cell.alignment = _LEFT

            val_cell = ws.cell(row=r_idx, column=2, value=value)
            val_cell.font      = Font(name="Calibri", size=11)
            val_cell.alignment = _LEFT

        title_cell = ws.cell(row=1, column=1,
                              value="📦  Current Stock Report — Summary")
        ws.merge_cells("A1:B1")
        title_cell.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill      = _GREEN_FILL
        title_cell.alignment = _CENTER
        ws.row_dimensions[1].height = 36
